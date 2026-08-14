#!/usr/bin/env python3
import argparse
import collections
import re
from urllib.parse import unquote, urlsplit

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RULES = [
    ("命令注入/恶意下载", "高", re.compile(
        r"(?:;|\||`|\$\(|%3[bB]|%7[cC])|(?:\brm\s*[-+]rf\b)|(?:\b(?:wget|curl|busybox|tftp)\b.*https?[:/])|"
        r"(?:/bin/(?:sh|bash))|(?:todo=syscmd)|(?:cmd=.*(?:wget|curl|sh\b|bash\b|chmod))|(?:mozi\.)", re.I)),
    ("路径穿越/文件读取", "高", re.compile(
        r"(?:\.\./|\.\.\\|%2e%2e|%252e%252e)|(?:/etc/(?:passwd|shadow|hosts))|"
        r"(?:/proc/self/(?:environ|cmdline|maps))|(?:file://)", re.I)),
    ("敏感文件/凭据探测", "高", re.compile(
        r"(?:^|/)(?:\.env(?:\.|/|$)|\.git(?:/|$)|\.svn(?:/|$)|\.hg(?:/|$)|id_rsa(?:$|[?]))|"
        r"(?:aws/credentials)|(?:\.docker/config\.json)|(?:config\.(?:php|yml|yaml|json)(?:$|[?]))|"
        r"(?:wp-config\.php)|(?:application\.properties)|(?:actuator/(?:env|heapdump|configprops))", re.I)),
    ("WebShell/RCE 漏洞扫描", "高", re.compile(
        r"(?:/cgi-bin/)|(?:setup\.cgi)|(?:shell\.php)|(?:cmd\.php)|(?:webshell)|(?:phpunit)|"
        r"(?:vendor/phpunit)|(?:invokefunction)|(?:call_user_func_array)|(?:eval-stdin\.php)|"
        r"(?:boaform/admin/formLogin)|(?:HNAP1)|(?:GponForm/diag_Form)|(?:goform/)|"
        r"(?:\.php\?(?:[^#]*)(?:cmd|exec|command)=)", re.I)),
    ("SQL 注入", "高", re.compile(
        r"(?:\bunion(?:\s|%20|\+)+(?:all(?:\s|%20|\+)+)?select\b)|(?:\bsleep\s*\()|"
        r"(?:\bbenchmark\s*\()|(?:\bwaitfor(?:\s|%20|\+)+delay\b)|(?:\bor(?:\s|%20|\+)+1=1\b)|"
        r"(?:information_schema)|(?:extractvalue\s*\()|(?:updatexml\s*\()", re.I)),
    ("XSS/脚本注入", "高", re.compile(
        r"(?:<script\b)|(?:javascript:)|(?:onerror\s*=)|(?:onload\s*=)|(?:%3cscript)|"
        r"(?:document\.cookie)|(?:alert\s*\()", re.I)),
    ("模板/表达式注入", "高", re.compile(
        r"(?:\$\{jndi:)|(?:\{\{.*(?:config|self|class|mro|subclasses).*(?:\}\}))|"
        r"(?:%7b%7b)|(?:_memberAccess)|(?:ognl)|(?:class\.module\.classLoader)", re.I)),
    ("CMS 漏洞扫描", "中", re.compile(
        r"(?:/wp-admin(?:/|$))|(?:/wp-login\.php)|(?:/wp-content/)|(?:/wp-includes/)|"
        r"(?:/xmlrpc\.php)|(?:/administrator(?:/|$))|(?:/joomla/)|(?:/drupal/)|"
        r"(?:/sites/default/files/)|(?:/typo3/)", re.I)),
    ("数据库/管理后台探测", "中", re.compile(
        r"(?:/phpmyadmin(?:/|$))|(?:/pma(?:/|$))|(?:/adminer(?:\.php|/|$))|"
        r"(?:/manager/html)|(?:/jmx-console)|(?:/solr/admin)|(?:/console/)|"
        r"(?:/server-status)|(?:/elasticsearch/)|(?:/_cat/)", re.I)),
    ("备份/源码泄露探测", "中", re.compile(
        r"(?:\.(?:bak|backup|old|orig|save|swp|sql|dump|tar|tgz|gz|zip|7z)(?:$|[?]))|"
        r"(?:/(?:backup|backups|dump|database|db)(?:/|\.|$))|(?:~(?:$|[?]))", re.I)),
    ("云元数据探测", "高", re.compile(
        r"(?:169\.254\.169\.254)|(?:metadata\.google\.internal)|(?:latest/meta-data)|"
        r"(?:computeMetadata/v1)", re.I)),
    ("通用漏洞扫描器特征", "中", re.compile(
        r"(?:/\.well-known/security\.txt\.php)|(?:/v2/_catalog)|(?:/swagger-ui(?:/|$))|"
        r"(?:/api-docs(?:/|$))|(?:/openapi\.json)|(?:/graphql\?query=.*__schema)", re.I)),
]

ILLEGAL_XML = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def clean(value):
    return ILLEGAL_XML.sub("", value) if isinstance(value, str) else value


def decode_repeated(value):
    decoded = value
    for _ in range(3):
        next_value = unquote(decoded.replace("+", " "))
        if next_value == decoded:
            break
        decoded = next_value
    return decoded


def classify(url):
    decoded = decode_repeated(url)
    matches = []
    for category, confidence, pattern in RULES:
        found = pattern.search(decoded)
        if found:
            evidence = found.group(0).replace("\n", " ")[:120]
            matches.append((category, confidence, evidence))
    if not matches:
        return None
    confidence = "高" if any(item[1] == "高" for item in matches) else "中"
    categories = "；".join(dict.fromkeys(item[0] for item in matches))
    evidence = "；".join(dict.fromkeys(item[2] for item in matches))
    return categories, confidence, evidence, decoded


def style(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="9C0006")
        cell.alignment = Alignment(horizontal="center")
    for column in ws.columns:
        sample = [str(cell.value or "") for cell in list(column)[:500]]
        width = min(max(max((len(v) for v in sample), default=8) + 2, 10), 65)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source")
    parser.add_argument("output")
    args = parser.parse_args()

    source = load_workbook(args.source, read_only=True, data_only=True)
    detail = source["请求明细"]
    headers = [cell.value for cell in next(detail.iter_rows(min_row=1, max_row=1))]
    index = {name: i for i, name in enumerate(headers)}

    malicious = []
    for values in detail.iter_rows(min_row=2, values_only=True):
        url = str(values[index["链接"]] or "")
        result = classify(url)
        if result:
            categories, confidence, evidence, decoded = result
            malicious.append((values, categories, confidence, evidence, decoded))

    grouped = collections.defaultdict(list)
    for values, categories, confidence, evidence, decoded in malicious:
        grouped[(values[index["链接"]], categories, confidence, evidence, decoded)].append(values)

    wb = Workbook()
    overview = wb.active
    overview.title = "概览"
    overview.append(["项目", "值"])
    overview.append(["原始请求数", detail.max_row - 1])
    overview.append(["疑似攻击请求数", len(malicious)])
    overview.append(["疑似攻击链接数", len(grouped)])
    overview.append(["筛选说明", "基于 URL 特征规则筛选；高置信度通常包含明确利用载荷，中置信度可能只是扫描或误报，需结合响应码复核。"])
    style(overview)

    summary = wb.create_sheet("攻击链接汇总")
    summary.append(["攻击分类", "置信度", "链接", "解码后链接", "命中特征", "请求次数", "独立IP数", "首次时间UTC", "最后时间UTC", "状态码分布"])
    for (url, categories, confidence, evidence, decoded), rows in sorted(grouped.items(), key=lambda item: -len(item[1])):
        status = collections.Counter(row[index["状态码"]] for row in rows)
        summary.append([clean(categories), confidence, clean(url), clean(decoded), clean(evidence), len(rows),
                        len({row[index["IP"]] for row in rows}),
                        min(row[index["时间UTC"]] for row in rows),
                        max(row[index["时间UTC"]] for row in rows),
                        ", ".join(f"{code}:{count}" for code, count in sorted(status.items()))])
    style(summary)

    requests = wb.create_sheet("攻击请求明细")
    requests.append(["攻击分类", "置信度", "命中特征", "解码后链接"] + headers)
    for values, categories, confidence, evidence, decoded in malicious:
        requests.append([clean(categories), confidence, clean(evidence), clean(decoded)] + [clean(value) for value in values])
    style(requests)

    categories_ws = wb.create_sheet("分类统计")
    categories_ws.append(["攻击分类", "请求次数", "独立链接数", "独立IP数"])
    for category in sorted({part for _, cats, _, _, _ in malicious for part in cats.split("；")}):
        items = [(values, cats) for values, cats, _, _, _ in malicious if category in cats.split("；")]
        categories_ws.append([category, len(items), len({v[index["链接"]] for v, _ in items}), len({v[index["IP"]] for v, _ in items})])
    style(categories_ws)

    wb.save(args.output)
    print(f"requests={len(malicious)} links={len(grouped)} output={args.output}")


if __name__ == "__main__":
    main()
