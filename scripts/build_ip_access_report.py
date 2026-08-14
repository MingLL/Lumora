#!/usr/bin/env python3
import argparse
import collections
import datetime as dt
import gzip
import ipaddress
import json
import re
import urllib.request
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OLD_RE = re.compile(
    r'^(?P<proxy>\S+) "(?P<ip>[^"]+)" \[(?P<time>[^]]+)] '
    r'"(?P<method>\S+) (?P<path>\S+) [^"]+" (?P<status>\d+) (?P<size>\d+) '
    r'"(?P<referer>[^"]*)" "(?P<ua>[^"]*)"$'
)


def iter_text(path):
    opener = gzip.open if path.suffix == ".gz" else open
    with opener(path, "rt", encoding="utf-8", errors="replace") as handle:
        yield from handle


def parse_logs(log_dir):
    rows = []
    cutoff = dt.date(2026, 8, 3)
    for path in sorted(log_dir.glob("access-20*.log")):
        for line in iter_text(path):
            match = OLD_RE.match(line.rstrip("\n"))
            if not match:
                continue
            item = match.groupdict()
            when = dt.datetime.strptime(item["time"], "%d/%b/%Y:%H:%M:%S %z")
            if when.date() >= cutoff:
                continue
            ip = item["ip"].split(",")[0].strip()
            rows.append({
                "time": when.astimezone(dt.timezone.utc).replace(tzinfo=None),
                "ip": ip,
                "method": item["method"],
                "scheme": "https",
                "host": "lumora.love",
                "path": item["path"],
                "status": int(item["status"]),
                "size": int(item["size"]),
                "referer": item["referer"],
                "ua": item["ua"],
                "source": "Lumora/Nginx",
            })

    for path in sorted(log_dir.glob("access.log*")):
        for line in iter_text(path):
            try:
                item = json.loads(line)
                when_text = item.get("StartUTC") or item.get("time")
                when = dt.datetime.fromisoformat(when_text.replace("Z", "+00:00"))
                ip = item.get("ClientHost") or item.get("ClientAddr", "").rsplit(":", 1)[0]
                if not ip:
                    continue
                rows.append({
                    "time": when.astimezone(dt.timezone.utc).replace(tzinfo=None),
                    "ip": ip,
                    "method": item.get("RequestMethod", ""),
                    "scheme": item.get("RequestScheme", ""),
                    "host": item.get("RequestHost", ""),
                    "path": item.get("RequestPath", ""),
                    "status": int(item.get("DownstreamStatus", 0)),
                    "size": int(item.get("DownstreamContentSize", 0)),
                    "referer": item.get("request_Referer", ""),
                    "ua": item.get("request_User-Agent", ""),
                    "source": "Traefik",
                })
            except (json.JSONDecodeError, TypeError, ValueError):
                continue

    unique = {}
    for row in rows:
        key = (row["time"], row["ip"], row["method"], row["host"], row["path"], row["status"])
        unique[key] = row
    return sorted(unique.values(), key=lambda row: row["time"])


def is_public(ip):
    try:
        obj = ipaddress.ip_address(ip)
        return obj.is_global
    except ValueError:
        return False


def load_geo(path):
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}


def query_geo(ips, cache_path):
    cache = load_geo(cache_path)
    pending = [ip for ip in sorted(ips) if is_public(ip) and ip not in cache]
    fields = "status,message,query,country,regionName,city,isp,org,as,proxy,hosting,mobile"
    for offset in range(0, len(pending), 100):
        batch = pending[offset:offset + 100]
        body = json.dumps([{"query": ip, "fields": fields, "lang": "zh-CN"} for ip in batch]).encode()
        request = urllib.request.Request(
            "http://ip-api.com/batch",
            data=body,
            headers={"Content-Type": "application/json", "User-Agent": "Lumora access audit/1.0"},
        )
        with urllib.request.urlopen(request, timeout=30) as response:
            for item in json.load(response):
                cache[item.get("query", "")] = item
        cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    for ip in ips:
        if ip not in cache:
            cache[ip] = {"status": "private", "query": ip, "country": "内网/保留地址"}
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    return cache


def query_geo_mmdb(ips, mmdb_path, cache_path):
    import maxminddb

    cache = load_geo(cache_path)
    with maxminddb.open_database(str(mmdb_path)) as reader:
        for ip in sorted(ips):
            if ip in cache:
                continue
            if not is_public(ip):
                cache[ip] = {"status": "private", "query": ip, "country": "内网/保留地址"}
                continue
            record = reader.get(ip) or {}
            names = lambda obj: (obj or {}).get("names", {})
            cache[ip] = {
                "status": "success" if record else "not-found",
                "query": ip,
                "country": names(record.get("country")).get("zh-CN") or names(record.get("country")).get("en", ""),
                "regionName": (names((record.get("subdivisions") or [{}])[0]).get("zh-CN")
                               or names((record.get("subdivisions") or [{}])[0]).get("en", "")),
                "city": names(record.get("city")).get("zh-CN") or names(record.get("city")).get("en", ""),
            }
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    return cache


def bot_label(ua):
    text = ua.lower()
    names = (("googlebot", "Googlebot"), ("bingbot", "Bingbot"), ("bytespider", "Bytespider"),
             ("bot", "其他爬虫"), ("spider", "其他爬虫"), ("crawler", "其他爬虫"),
             ("zgrab", "扫描器"), ("masscan", "扫描器"), ("curl", "脚本/命令行"),
             ("python", "脚本/命令行"))
    return next((label for token, label in names if token in text), "普通/未知")


def location(geo):
    parts = [geo.get("country", ""), geo.get("regionName", ""), geo.get("city", "")]
    return " / ".join(part for part in parts if part)


def style_sheet(ws, freeze="A2", autofilter=True):
    ws.freeze_panes = freeze
    if autofilter and ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for column in ws.columns:
        values = [str(cell.value or "") for cell in list(column)[:500]]
        width = min(max(max(map(len, values), default=8) + 2, 10), 55)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def build_excel(rows, geo, output):
    wb = Workbook()
    overview = wb.active
    overview.title = "概览"
    start = min(row["time"] for row in rows)
    end = max(row["time"] for row in rows)
    unique_ips = {row["ip"] for row in rows}
    overview.append(["项目", "值"])
    overview.append(["统计范围（UTC）", f"{start:%Y-%m-%d %H:%M:%S} 至 {end:%Y-%m-%d %H:%M:%S}"])
    overview.append(["日志覆盖说明", "服务器仅保留 2026-07-27 起的日志，未达到完整一个月；2026-08-03 起以 Traefik 入口日志为准。"])
    overview.append(["IP 地点数据", "DB-IP City Lite 2026-07，CC BY 4.0；地点为 IP 网络出口的近似位置，不代表访客精确位置。"])
    overview.append(["总请求数", len(rows)])
    overview.append(["独立 IP 数", len(unique_ips)])
    overview.append(["成功请求（2xx/3xx）", sum(200 <= row["status"] < 400 for row in rows)])
    overview.append(["客户端错误（4xx）", sum(400 <= row["status"] < 500 for row in rows)])
    overview.append(["服务端错误（5xx）", sum(row["status"] >= 500 for row in rows)])
    style_sheet(overview, autofilter=False)

    by_ip = collections.defaultdict(list)
    for row in rows:
        by_ip[row["ip"]].append(row)
    ws = wb.create_sheet("IP汇总")
    ws.append(["IP", "请求次数", "首次访问UTC", "最后访问UTC", "独立链接数", "2xx", "3xx", "4xx", "5xx", "类型", "地点", "运营商/ISP", "组织", "ASN", "代理", "托管机房"])
    for ip, items in sorted(by_ip.items(), key=lambda pair: (-len(pair[1]), pair[0])):
        info = geo.get(ip, {})
        uas = collections.Counter(bot_label(item["ua"]) for item in items)
        ws.append([ip, len(items), min(x["time"] for x in items), max(x["time"] for x in items),
                   len({(x["host"], x["path"]) for x in items}),
                   sum(200 <= x["status"] < 300 for x in items), sum(300 <= x["status"] < 400 for x in items),
                   sum(400 <= x["status"] < 500 for x in items), sum(x["status"] >= 500 for x in items),
                   uas.most_common(1)[0][0], location(info), info.get("isp", ""), info.get("org", ""),
                   info.get("as", ""), "是" if info.get("proxy") else "", "是" if info.get("hosting") else ""])
    style_sheet(ws)

    by_ip_url = collections.defaultdict(list)
    by_url = collections.defaultdict(list)
    for row in rows:
        url = f'{row["scheme"]}://{row["host"]}{row["path"]}'
        by_ip_url[(row["ip"], url)].append(row)
        by_url[url].append(row)
    ws = wb.create_sheet("IP与链接")
    ws.append(["IP", "链接", "请求次数", "首次访问UTC", "最后访问UTC", "主要状态码", "地点", "主要User-Agent类型"])
    for (ip, url), items in sorted(by_ip_url.items(), key=lambda pair: (-len(pair[1]), pair[0])):
        statuses = collections.Counter(item["status"] for item in items)
        uas = collections.Counter(bot_label(item["ua"]) for item in items)
        ws.append([ip, url, len(items), min(x["time"] for x in items), max(x["time"] for x in items),
                   statuses.most_common(1)[0][0], location(geo.get(ip, {})), uas.most_common(1)[0][0]])
    style_sheet(ws)

    ws = wb.create_sheet("链接汇总")
    ws.append(["链接", "请求次数", "独立IP数", "2xx", "3xx", "4xx", "5xx", "首次访问UTC", "最后访问UTC"])
    for url, items in sorted(by_url.items(), key=lambda pair: (-len(pair[1]), pair[0])):
        ws.append([url, len(items), len({x["ip"] for x in items}),
                   sum(200 <= x["status"] < 300 for x in items), sum(300 <= x["status"] < 400 for x in items),
                   sum(400 <= x["status"] < 500 for x in items), sum(x["status"] >= 500 for x in items),
                   min(x["time"] for x in items), max(x["time"] for x in items)])
    style_sheet(ws)

    ws = wb.create_sheet("请求明细")
    ws.append(["时间UTC", "IP", "地点", "方法", "链接", "状态码", "响应字节", "类型", "Referer", "User-Agent", "日志来源"])
    for row in rows:
        url = f'{row["scheme"]}://{row["host"]}{row["path"]}'
        ws.append([row["time"], row["ip"], location(geo.get(row["ip"], {})), row["method"], url,
                   row["status"], row["size"], bot_label(row["ua"]), row["referer"], row["ua"], row["source"]])
    style_sheet(ws)
    wb.save(output)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("log_dir", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--geo-cache", type=Path, required=True)
    parser.add_argument("--mmdb", type=Path)
    parser.add_argument("--count-only", action="store_true")
    args = parser.parse_args()
    rows = parse_logs(args.log_dir)
    ips = {row["ip"] for row in rows}
    print(json.dumps({"requests": len(rows), "unique_ips": len(ips), "start": min(x["time"] for x in rows).isoformat(), "end": max(x["time"] for x in rows).isoformat()}, ensure_ascii=False))
    if args.count_only:
        return
    geo = query_geo_mmdb(ips, args.mmdb, args.geo_cache) if args.mmdb else query_geo(ips, args.geo_cache)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    build_excel(rows, geo, args.output)
    print(args.output)


if __name__ == "__main__":
    main()
